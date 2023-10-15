from typing import Iterator

from openpyxl import load_workbook

from configs import config
from schemas import RowValues


class ExcelReader:
    workbook = load_workbook(config.excel_file_path)
    sheet = workbook.active

    def read_rows(self) -> Iterator[RowValues]:
        """
        Возвращает строку таблицы. Строку с заголовком(1 строку) пропускает.
        :return:
        """
        for index, row in enumerate(self.sheet.iter_rows()):
            if index == 0:
                continue

            yield RowValues(
                id_tovar=row[0].value,
                tovar=row[1].value,
                id_isg=row[2].value,
                isg=row[3].value,
                country=row[4].value,
                barcode=row[5].value,
            )
